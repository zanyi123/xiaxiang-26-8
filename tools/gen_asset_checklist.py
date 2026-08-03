# -*- coding: utf-8 -*-
"""生成项目素材总清单 Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "素材总清单"

# ===== 样式定义 =====
title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
content_font = Font(name="微软雅黑", size=10)
category_font = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")

title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
img_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
mdl_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
vid_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
aud_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ===== 列宽（格子大点）=====
col_widths = [12, 10, 22, 28, 35, 14, 30, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ===== 行高 =====
ws.row_dimensions[1].height = 40  # 标题行
ws.row_dimensions[2].height = 28  # 表头行

# ===== 标题行 =====
ws.merge_cells("A1:H1")
cell = ws["A1"]
cell.value = "薪火侨乡项目 — 后台素材上传总清单"
cell.font = title_font
cell.fill = title_fill
cell.alignment = center_align

# ===== 表头 =====
headers = ["Slot编号", "类型", "所属模块", "YAML字段路径", "用途说明", "格式要求", "COS存储路径", "上传状态"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border

# ===== 数据 =====
data = [
    # 模块1: 云游侨乡 - 建筑 (3栋建筑)
    # 建筑1: 开平碉楼
    ("IMG-01-01", "图片", "云游侨乡-建筑总览", "buildings[0].coverImage", "开平碉楼 封面图", "jpg/png", "根目录 cover-kaiping.jpg", "☐ 待上传"),
    ("MDL-01-01", "3D模型", "云游侨乡-建筑总览", "buildings[0].modelKey", "开平碉楼 3D高斯溅射模型", ".splat / .ply / .glb", "models/kaiping.splat", "☐ 待上传"),
    ("VID-01-01", "视频", "云游侨乡-建筑总览", "buildings[0].videoKey", "开平碉楼 4K宣传片", ".mp4 / .mov / .m4v", "videos/kaiping-4k.mp4", "☐ 待上传"),
    # 建筑2: 自力村碉楼群
    ("IMG-01-02", "图片", "云游侨乡-建筑总览", "buildings[1].coverImage", "自力村碉楼群 封面图", "jpg/png", "根目录 cover-zili.jpg", "☐ 待上传"),
    ("MDL-01-02", "3D模型", "云游侨乡-建筑总览", "buildings[1].modelKey", "自力村碉楼群 3D模型", ".splat / .ply / .glb", "models/zili.splat", "☐ 待上传"),
    ("VID-01-02", "视频", "云游侨乡-建筑总览", "buildings[1].videoKey", "自力村碉楼群 4K宣传片", ".mp4 / .mov / .m4v", "videos/zili-4k.mp4", "☐ 待上传"),
    # 建筑3: 赤坎古镇骑楼
    ("IMG-01-03", "图片", "云游侨乡-建筑总览", "buildings[2].coverImage", "赤坎古镇骑楼 封面图", "jpg/png", "根目录 cover-chikan.jpg", "☐ 待上传"),
    ("MDL-01-03", "3D模型", "云游侨乡-建筑总览", "buildings[2].modelKey", "赤坎古镇骑楼 3D模型", ".splat / .ply / .glb", "models/chikan.splat", "☐ 待上传"),
    ("VID-01-03", "视频", "云游侨乡-建筑总览", "buildings[2].videoKey", "赤坎古镇骑楼 4K宣传片", ".mp4 / .mov / .m4v", "videos/chikan-4k.mp4", "☐ 待上传"),

    # 模块2: 云游侨乡 - 地点 (6个地点)
    ("IMG-02-01", "图片", "云游侨乡-地点", "locations[0].imageKey", "宝源坊碉楼群 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-02-01", "3D模型", "云游侨乡-地点", "locations[0].modelKey", "宝源坊碉楼群 3D模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-02-02", "图片", "云游侨乡-地点", "locations[1].imageKey", "方形古井 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-02-02", "3D模型", "云游侨乡-地点", "locations[1].modelKey", "方形古井 3D模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-02-03", "图片", "云游侨乡-地点", "locations[2].imageKey", "青石板路 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-02-03", "3D模型", "云游侨乡-地点", "locations[2].modelKey", "青石板路 3D模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-02-04", "图片", "云游侨乡-地点", "locations[3].imageKey", "古榕树群 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-02-04", "3D模型", "云游侨乡-地点", "locations[3].modelKey", "古榕树群 3D模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-02-05", "图片", "云游侨乡-地点", "locations[4].imageKey", "陈氏宗祠 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-02-05", "3D模型", "云游侨乡-地点", "locations[4].modelKey", "陈氏宗祠 3D模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-02-06", "图片", "云游侨乡-地点", "locations[5].imageKey", "侨批馆 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-02-06", "3D模型", "云游侨乡-地点", "locations[5].modelKey", "侨批馆 3D模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),

    # 模块4: 侨乡故事 (4个故事)
    ("IMG-04-01", "图片", "侨乡故事", "stories[0].coverImage", "故事1 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("AUD-04-01", "音频", "侨乡故事", "stories[0].audioKey", "故事1 朗读音频", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("IMG-04-02", "图片", "侨乡故事", "stories[1].coverImage", "故事2 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("AUD-04-02", "音频", "侨乡故事", "stories[1].audioKey", "故事2 朗读音频", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("IMG-04-03", "图片", "侨乡故事", "stories[2].coverImage", "故事3 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("AUD-04-03", "音频", "侨乡故事", "stories[2].audioKey", "故事3 朗读音频", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("IMG-04-04", "图片", "侨乡故事", "stories[3].coverImage", "故事4 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("AUD-04-04", "音频", "侨乡故事", "stories[3].audioKey", "故事4 朗读音频", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),

    # 模块5: 建筑解剖 (5个部位)
    ("IMG-05-01", "图片", "建筑解剖", "anatomies[0].imageKey", "屋顶 细节图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-05-01", "3D模型", "建筑解剖", "anatomies[0].modelKey", "屋顶 分体模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-05-02", "图片", "建筑解剖", "anatomies[1].imageKey", "八角窗 细节图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-05-02", "3D模型", "建筑解剖", "anatomies[1].modelKey", "八角窗 分体模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-05-03", "图片", "建筑解剖", "anatomies[2].imageKey", "燕子窝 细节图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-05-03", "3D模型", "建筑解剖", "anatomies[2].modelKey", "燕子窝 分体模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-05-04", "图片", "建筑解剖", "anatomies[3].imageKey", "青砖墙体 细节图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-05-04", "3D模型", "建筑解剖", "anatomies[3].modelKey", "青砖墙体 分体模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),
    ("IMG-05-05", "图片", "建筑解剖", "anatomies[4].imageKey", "灰塑装饰 细节图", "jpg/png", "images/", "☐ 待上传"),
    ("MDL-05-05", "3D模型", "建筑解剖", "anatomies[4].modelKey", "灰塑装饰 分体模型", ".splat / .ply / .glb", "models/", "☐ 待上传"),

    # 模块6: 知识库 (6个条目)
    ("IMG-06-01", "图片", "知识库", "knowledge[0].coverImage", "知识条目1 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-06-02", "图片", "知识库", "knowledge[1].coverImage", "知识条目2 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-06-03", "图片", "知识库", "knowledge[2].coverImage", "知识条目3 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-06-04", "图片", "知识库", "knowledge[3].coverImage", "知识条目4 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-06-05", "图片", "知识库", "knowledge[4].coverImage", "知识条目5 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-06-06", "图片", "知识库", "knowledge[5].coverImage", "知识条目6 封面图", "jpg/png", "images/", "☐ 待上传"),

    # 模块7: 民俗文化 (6项)
    ("IMG-07-01", "图片", "民俗文化", "cultures[0].coverImage", "民俗1 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-07-02", "图片", "民俗文化", "cultures[1].coverImage", "民俗2 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-07-03", "图片", "民俗文化", "cultures[2].coverImage", "民俗3 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-07-04", "图片", "民俗文化", "cultures[3].coverImage", "民俗4 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-07-05", "图片", "民俗文化", "cultures[4].coverImage", "民俗5 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-07-06", "图片", "民俗文化", "cultures[5].coverImage", "民俗6 封面图", "jpg/png", "images/", "☐ 待上传"),

    # 模块8: 实践日志 (4篇)
    ("IMG-08-01", "图片", "实践日志", "blogPosts[0].coverImage", "日志1 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-08-02", "图片", "实践日志", "blogPosts[1].coverImage", "日志2 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-08-03", "图片", "实践日志", "blogPosts[2].coverImage", "日志3 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-08-04", "图片", "实践日志", "blogPosts[3].coverImage", "日志4 封面图", "jpg/png", "images/", "☐ 待上传"),

    # 模块9: 视频展播 (4个视频)
    ("IMG-09-01", "图片", "视频展播", "videos[0].coverImage", "视频1 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("VID-09-01", "视频", "视频展播", "videos[0].videoKey", "视频1 正片", ".mp4 / .mov / .m4v", "videos/", "☐ 待上传"),
    ("IMG-09-02", "图片", "视频展播", "videos[1].coverImage", "视频2 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("VID-09-02", "视频", "视频展播", "videos[1].videoKey", "视频2 正片", ".mp4 / .mov / .m4v", "videos/", "☐ 待上传"),
    ("IMG-09-03", "图片", "视频展播", "videos[2].coverImage", "视频3 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("VID-09-03", "视频", "视频展播", "videos[2].videoKey", "视频3 正片", ".mp4 / .mov / .m4v", "videos/", "☐ 待上传"),
    ("IMG-09-04", "图片", "视频展播", "videos[3].coverImage", "视频4 封面图", "jpg/png", "images/", "☐ 待上传"),
    ("VID-09-04", "视频", "视频展播", "videos[3].videoKey", "视频4 正片", ".mp4 / .mov / .m4v", "videos/", "☐ 待上传"),

    # 模块10: 团队成员 (6人)
    ("IMG-10-01", "图片", "团队成员", "team[0].avatar", "成员1 头像（建议正方形）", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-10-02", "图片", "团队成员", "team[1].avatar", "成员2 头像（建议正方形）", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-10-03", "图片", "团队成员", "team[2].avatar", "成员3 头像（建议正方形）", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-10-04", "图片", "团队成员", "team[3].avatar", "成员4 头像（建议正方形）", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-10-05", "图片", "团队成员", "team[4].avatar", "成员5 头像（建议正方形）", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-10-06", "图片", "团队成员", "team[5].avatar", "成员6 头像（建议正方形）", "jpg/png", "images/", "☐ 待上传"),

    # 模块11: 老照片对比 (4组×2=8张)
    ("IMG-11-01", "图片", "老照片对比", "photoCompares[0].oldImageKey", "对比1 历史老照片", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-11-02", "图片", "老照片对比", "photoCompares[0].newImageKey", "对比1 现状新照片", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-11-03", "图片", "老照片对比", "photoCompares[1].oldImageKey", "对比2 历史老照片", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-11-04", "图片", "老照片对比", "photoCompares[1].newImageKey", "对比2 现状新照片", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-11-05", "图片", "老照片对比", "photoCompares[2].oldImageKey", "对比3 历史老照片", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-11-06", "图片", "老照片对比", "photoCompares[2].newImageKey", "对比3 现状新照片", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-11-07", "图片", "老照片对比", "photoCompares[3].oldImageKey", "对比4 历史老照片", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-11-08", "图片", "老照片对比", "photoCompares[3].newImageKey", "对比4 现状新照片", "jpg/png", "images/", "☐ 待上传"),

    # 模块12: 侨批文化 (5封)
    ("IMG-12-01", "图片", "侨批文化", "qiaopi[0].imageKey", "侨批1 扫描件", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-12-02", "图片", "侨批文化", "qiaopi[1].imageKey", "侨批2 扫描件", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-12-03", "图片", "侨批文化", "qiaopi[2].imageKey", "侨批3 扫描件", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-12-04", "图片", "侨批文化", "qiaopi[3].imageKey", "侨批4 扫描件", "jpg/png", "images/", "☐ 待上传"),
    ("IMG-12-05", "图片", "侨批文化", "qiaopi[4].imageKey", "侨批5 扫描件", "jpg/png", "images/", "☐ 待上传"),

    # 模块13: 方言学习 (8条)
    ("AUD-13-01", "音频", "方言学习", "dialects[0].audioKey", "方言1 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("AUD-13-02", "音频", "方言学习", "dialects[1].audioKey", "方言2 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("AUD-13-03", "音频", "方言学习", "dialects[2].audioKey", "方言3 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("AUD-13-04", "音频", "方言学习", "dialects[3].audioKey", "方言4 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("AUD-13-05", "音频", "方言学习", "dialects[4].audioKey", "方言5 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("AUD-13-06", "音频", "方言学习", "dialects[5].audioKey", "方言6 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("AUD-13-07", "音频", "方言学习", "dialects[6].audioKey", "方言7 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),
    ("AUD-13-08", "音频", "方言学习", "dialects[7].audioKey", "方言8 录音", "mp3/wav/aac/m4a", "audio/", "☐ 待上传"),

    # 模块15: 虚拟盖章 (8枚)
    ("IMG-15-01", "图片", "虚拟盖章", "stamps[0].imageKey", "印章1 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
    ("IMG-15-02", "图片", "虚拟盖章", "stamps[1].imageKey", "印章2 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
    ("IMG-15-03", "图片", "虚拟盖章", "stamps[2].imageKey", "印章3 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
    ("IMG-15-04", "图片", "虚拟盖章", "stamps[3].imageKey", "印章4 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
    ("IMG-15-05", "图片", "虚拟盖章", "stamps[4].imageKey", "印章5 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
    ("IMG-15-06", "图片", "虚拟盖章", "stamps[5].imageKey", "印章6 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
    ("IMG-15-07", "图片", "虚拟盖章", "stamps[6].imageKey", "印章7 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
    ("IMG-15-08", "图片", "虚拟盖章", "stamps[7].imageKey", "印章8 图标（建议透明PNG）", "png", "images/", "☐ 待上传"),
]

# ===== 写入数据 =====
row = 3
for item in data:
    slot_id, typ, module, yaml_path, desc, fmt, cos_path, status = item

    # 类型颜色
    if typ == "图片":
        row_fill = img_fill
    elif typ == "3D模型":
        row_fill = mdl_fill
    elif typ == "视频":
        row_fill = vid_fill
    elif typ == "音频":
        row_fill = aud_fill
    else:
        row_fill = None

    for col, val in enumerate(item, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = content_font
        c.alignment = center_align if col <= 2 else left_align
        c.border = thin_border
        if row_fill:
            c.fill = row_fill

    ws.row_dimensions[row].height = 25
    row += 1

# ===== 冻结表头 =====
ws.freeze_panes = "A3"

# ===== 添加统计Sheet =====
ws2 = wb.create_sheet("统计汇总")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 50

ws2.merge_cells("A1:C1")
c = ws2["A1"]
c.value = "素材统计汇总"
c.font = title_font
c.fill = title_fill
c.alignment = center_align
ws2.row_dimensions[1].height = 40

stats_headers = ["素材类型", "数量", "格式要求"]
for col, h in enumerate(stats_headers, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border
ws2.row_dimensions[2].height = 28

stats = [
    ("图片 (IMG)", sum(1 for d in data if d[1] == "图片"), "jpg / jpeg / png / gif / webp / bmp"),
    ("3D模型 (MDL)", sum(1 for d in data if d[1] == "3D模型"), ".splat / .ply / .obj / .glb / .gltf / .fbx"),
    ("视频 (VID)", sum(1 for d in data if d[1] == "视频"), ".mp4 / .mov / .m4v"),
    ("音频 (AUD)", sum(1 for d in data if d[1] == "音频"), ".mp3 / .wav / .aac / .m4a"),
]

for i, (name, count, fmt) in enumerate(stats, 3):
    ws2.cell(row=i, column=1, value=name).font = content_font
    ws2.cell(row=i, column=2, value=count).font = Font(name="微软雅黑", size=12, bold=True)
    ws2.cell(row=i, column=3, value=fmt).font = content_font
    for col in range(1, 4):
        ws2.cell(row=i, column=col).alignment = center_align
        ws2.cell(row=i, column=col).border = thin_border
    ws2.row_dimensions[i].height = 25

# 合计行
total_row = 3 + len(stats)
ws2.cell(row=total_row, column=1, value="合计").font = Font(name="微软雅黑", size=11, bold=True)
ws2.cell(row=total_row, column=2, value=len(data)).font = Font(name="微软雅黑", size=12, bold=True, color="FF0000")
ws2.cell(row=total_row, column=3, value="").font = content_font
for col in range(1, 4):
    ws2.cell(row=total_row, column=col).alignment = center_align
    ws2.cell(row=total_row, column=col).border = thin_border
    ws2.cell(row=total_row, column=col).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
ws2.row_dimensions[total_row].height = 28

# 保存
output_path = r"d:\JAVA\xiaxiang\docs\素材上传总清单.xlsx"
wb.save(output_path)
print(f"Excel 已生成: {output_path}")
print(f"素材总数: {len(data)} 项")
