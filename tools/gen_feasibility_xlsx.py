"""
生成项目可行性分析对照表 xlsx
基于 YAML 配置 + COS 静态资源架构评估项目树各内容
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 输出路径
out_path = os.path.join(os.path.dirname(__file__), "..", "项目可行性分析对照表.xlsx")
out_path = os.path.abspath(out_path)

wb = Workbook()
ws = wb.active
ws.title = "可行性分析"

# ============ 样式定义 ============
title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
cell_font = Font(name="微软雅黑", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 判定颜色填充
fill_ok = PatternFill("solid", fgColor="5CB85C")       # 绿色 能做
fill_swap = PatternFill("solid", fgColor="F0AD4E")     # 橙色 换方式
fill_drop = PatternFill("solid", fgColor="D9534F")     # 红色 舍
fill_header = PatternFill("solid", fgColor="2C3E50")   # 深蓝 表头
fill_title = PatternFill("solid", fgColor="1A1A2E")    # 深黑 标题
fill_alt = PatternFill("solid", fgColor="F8F9FA")      # 浅灰 隔行

thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============ 数据 ============
title = "云游侨乡·宝源坊 — 项目可行性分析对照表（基于 YAML 配置 + COS 静态资源架构）"

headers = ["项目树内容", "判定", "理由 / 替代方案"]

rows = [
    # 首屏
    ("【01 首屏 · 遇见宝源坊】", "", ""),
    ("首屏视频背景", "✅ 能做", "COS 存视频，<video> 标签自动播放即可，与现有架构完全匹配"),
    ("项目标题 / 核心文案", "✅ 能做", "纯 HTML/CSS，写死在模板里"),
    ("云游宝源坊 主入口", "✅ 能做", "点击跳转到 /map 地图页"),
    ("我的宝源坊人生 次入口", "✅ 能做", "点击跳转到互动体验区块，纯前端 JS 实现"),

    # 云游
    ("【02 云游宝源坊】", "", ""),
    ("2.1 宝源坊3D全景（整村漫游）", "❌ 舍", "整村 3DGS 模型文件 1GB+，浏览器加载不了；整村漫游交互逻辑也是独立项目级工作量"),
    ("2.1 替代方案：2D航拍地图+建筑热点", "⚠️ 换方式", "用航拍图做底图，标注 6 个建筑点位，点击进入单栋 3D 模型。即现有 /map + /building/{id} 逻辑"),
    ("2.2 特色建筑展示（6栋卡片）", "✅ 能做", "YAML 配置建筑数据 + COS 存封面图，Thymeleaf 循环渲染"),
    ("2.3 新手引导三步流程", "✅ 能做", "纯 CSS 流程图，写死在模板里"),

    # 互动
    ("【03 我的宝源坊人生】", "", ""),
    ("3.1 出生选择（三选一）", "✅ 能做", "纯前端 JS，点击卡片记录选择"),
    ("3.2 第一阶段抉择", "✅ 能做", "纯前端 JS 状态管理"),
    ("3.3 第二阶段抉择", "✅ 能做", "纯前端 JS 状态管理"),
    ("3.4 9种人生结局映射", "✅ 能做", "YAML 配置 9 种结局 + 对应特色建筑，前端按 3 次选择查表"),
    ("3.5 分享海报生成", "⚠️ 换方式", "前端 html2canvas 生成图片，用户手动保存分享。不用后端生成"),
    ("3.5 跨设备查看人生结果", "❌ 舍", "没数据库存不了用户数据。改为 URL 参数编码 /life-result?o=1&c1=2&c2=3，他人点链接前端还原"),
    ("3.5 人生标签 / 建筑介绍", "✅ 能做", "结局对应的建筑信息从 YAML 读取"),

    # 建筑故事
    ("【04 建筑故事】", "", ""),
    ("建筑列表（六栋卡片）", "✅ 能做", "YAML buildings 列表 + COS 封面图"),
    ("单栋 3D 建筑模型", "✅ 能做", "COS 存 .splat 文件，现有 building-3d.js 已实现"),
    ("建筑影像-航拍/外观/细节（多图）", "⚠️ 换方式", "YAML Building 增加 images 列表字段，每项含 key+caption。模板改轮播/网格展示，CosService 已支持"),
    ("建筑故事-年代/背景/价值/特色", "✅ 能做", "YAML 加 story / era / background 等文字字段"),
    ("延伸-后人访谈", "❌ 舍", "无素材。拍摄+剪辑+转录是纪录片团队工作量，不是网站能解决的"),
    ("延伸-老照片对比", "❌ 舍", "缺百年前老照片素材。有素材再做，现在先跳过"),
    ("延伸-口述记忆", "⚠️ 换方式", "改为一段口述文字 + 一段音频（COS 存 mp3），<audio> 播放。不做视频交互"),

    # 认识 + 成果
    ("【05 认识宝源坊 / 项目成果】", "", ""),
    ("三句话读懂宝源坊", "✅ 能做", "纯 HTML/CSS 三列卡片"),
    ("宝源坊简介文字", "✅ 能做", "写死在模板里，或放 YAML"),
    ("项目背景（百千万工程等）", "⚠️ 换方式", "写死在 about.html 模板里，不放 YAML。内容不会变动"),
    ("实践过程（入村/航拍/建模/团队）", "⚠️ 换方式", "同上，模板写死。只有数字成果（建模数量等）放 YAML 方便改"),
    ("数字技术展示", "⚠️ 换方式", "模板写死，配示意图"),
    ("成果展示（模型/调研/视频/合作）", "⚠️ 换方式", "模板写死 + 部分数据从 YAML 读取"),
    ("合作单位信息", "✅ 能做", "写死在页脚模板里"),

    # 后台
    ("【06 后台管理】", "", ""),
    ("建筑/图片/视频/3D模型 内容管理", "❌ 舍", "不做传统后台。改为本地 Python 脚本批量上传 COS，输出 URL 填回 application.yml"),
    ("用户数据存储", "❌ 舍", "无数据库，不存用户数据。互动结果走 URL 参数"),
]

# ============ 写入数据 ============
# 标题行
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
title_cell = ws.cell(row=1, column=1, value=title)
title_cell.font = title_font
title_cell.alignment = center
title_cell.fill = fill_title
ws.row_dimensions[1].height = 40

# 表头行
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font = header_font
    c.alignment = center
    c.fill = fill_header
    c.border = border
ws.row_dimensions[2].height = 32

# 数据行
for i, (content, verdict, reason) in enumerate(rows, start=3):
    is_section = (verdict == "" and reason == "" and content != "")
    is_last = (i == len(rows) + 2)

    c1 = ws.cell(row=i, column=1, value=content)
    c2 = ws.cell(row=i, column=2, value=verdict)
    c3 = ws.cell(row=i, column=3, value=reason)

    for c in (c1, c2, c3):
        c.font = cell_font
        c.border = border
        c.alignment = left_wrap if c is c3 else center
        if is_section:
            c.font = Font(name="微软雅黑", size=11, bold=True, color="2C3E50")
            c.fill = PatternFill("solid", fgColor="E8F4F8")

    # 判定列颜色
    if "能做" in verdict:
        c2.fill = fill_ok
        c2.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    elif "换方式" in verdict:
        c2.fill = fill_swap
        c2.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    elif "舍" in verdict:
        c2.fill = fill_drop
        c2.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

    # 行高：分区标题行矮一点，数据行根据内容长度
    if is_section:
        ws.row_dimensions[i].height = 28
    else:
        # 根据理由文字长度估算行高
        reason_len = len(reason) if reason else 0
        height = max(40, min(90, 30 + reason_len // 8 * 6))
        ws.row_dimensions[i].height = height

# ============ 列宽 ============
ws.column_dimensions["A"].width = 38   # 项目树内容
ws.column_dimensions["B"].width = 14   # 判定
ws.column_dimensions["C"].width = 70   # 理由

# 冻结表头
ws.freeze_panes = "A3"

# ============ 第二个 Sheet：架构原则说明 ============
ws2 = wb.create_sheet("架构原则")

ws2.merge_cells("A1:B1")
t = ws2.cell(row=1, column=1, value="架构原则：YAML 配置 + COS 静态资源（不引入数据库）")
t.font = title_font
t.alignment = center
t.fill = fill_title
ws2.row_dimensions[1].height = 40

principles = [
    ("内容类型", "存储与更新方式"),
    ("建筑信息、故事文案、档案数据", "application.yml 配置，直接编辑 yml 重启生效"),
    ("图片、视频、3D 模型 (.splat)", "腾讯云 COS，本地 Python 脚本批量上传，输出 URL 填回 yml"),
    ("互动选择结果", "URL 参数编码，不存数据库"),
    ("项目成果/背景文字", "写死在 about.html 模板里，不放 YAML"),
    ("", ""),
    ("不做传统后台的理由", ""),
    ("1. 无数据库依赖", "pom.xml 无 mysql/mybatis/jpa，引入后台等于重构整个数据层"),
    ("2. 低频变动内容", "建筑信息、故事文案一学期改一次，yml 完全够用"),
    ("3. 素材上传更轻量", "本地脚本上传 COS 比后台管理系统更快，不用登录腾讯云控制台"),
    ("4. 避免过度设计", "三下乡展示项目保持架构纯净，不堆砌技术"),
]

for r, (k, v) in enumerate(principles, start=2):
    c1 = ws2.cell(row=r, column=1, value=k)
    c2 = ws2.cell(row=r, column=2, value=v)
    for c in (c1, c2):
        c.font = cell_font
        c.alignment = left_wrap
        c.border = border
    if r == 2:  # 表头
        for c in (c1, c2):
            c.font = header_font
            c.fill = fill_header
            c.alignment = center
    elif k.startswith("不做"):
        c1.font = Font(name="微软雅黑", size=11, bold=True, color="2C3E50")
        c1.fill = PatternFill("solid", fgColor="E8F4F8")
        c2.fill = PatternFill("solid", fgColor="E8F4F8")
    ws2.row_dimensions[r].height = 36

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 75

# ============ 保存 ============
wb.save(out_path)
print(f"✅ 已生成：{out_path}")
print(f"   Sheet1: 可行性分析（{len(rows)} 行数据）")
print(f"   Sheet2: 架构原则")
